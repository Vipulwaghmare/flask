from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime

class HandleExcel:
  def __init__(self, file_path, working_sheet = None):
    self.file_path = file_path
    self.wb = load_workbook(file_path)
    if working_sheet is None:
      # sets default sheet as working sheet
      self.working_sheet = self.wb.active 
    else:
      self.set_active_sheet(working_sheet)

  # To avoid writing save function and passing file name repeatedly
  def save_file(self):
    try:
      self.wb.save(self.file_path)
    except Exception as e:
      raise e

  # SHEETS OPERATIONS
  def get_sheet_names(self):
    return self.wb.sheetnames

  def get_active_sheet(self):
    return self.working_sheet
  
  def set_active_sheet(self, sheet_name):
    try: 
      self.working_sheet = self.wb[sheet_name] 
      return self.working_sheet
    except Exception as e:
      raise e
  
  def create_sheet(self, name):
    try:
      self.wb.create_sheet(name)
      self.save_file()
    except Exception as e:
      raise e

  # CELL OPERATIONS
  def get_value_of_cell(self, cell): 
    value = self.working_sheet[cell].value
    print("VALUE OF " + cell + " = " + value)
    return value
  
  def set_value_of_cell(self, cell, value):
    try:
      self.working_sheet[cell].value = value
      self.save_file()
    except Exception as e:
      raise e
  
  # Adds data to newest row
  # data is an array of data to be added
  # Appends data to existing data already available
  def append_data(self, data):
    try:
      self.working_sheet.append(data)
      self.save_file()
    except Exception as e:
      raise e
  
  # Accessing rows
  def get_data(self, row_end, column_end, row_start = 1, col_start = 1):
    response = [] # TO store response
    titles_array = []  # to store titles
    # For loop doesn't count the last one
    for row in range(row_start, row_end):
      row_object = {}
      for col in range(col_start, column_end):
        char = get_column_letter(col) # gets character A, B, C.. Z , etc 
        value = self.working_sheet[char + str(row)].value # A1, A2, getting values
        # INSERT TITLES ARRAY for 1st row
        if row == 1:
          titles_array.append(value)
        # CREATE DATA for rows other than first
        else:
          if (len(titles_array) > col): 
            if titles_array[col - 1] is not None:
              row_object[titles_array[col - 1]] = value
      
      # Inserting data except for 1st row
      if row != 1:
        response.append(row_object)
    return response
  
  def insert_data(self, data, new_sheet_name = None):
    # [{'Name': 'MEOW', 'Gender': 'Male', 'DOB': datetime.datetime(1998, 2, 22, 0, 0)}, {'Name': 'NHAU', 'Gender': 'new value', 'DOB': datetime.datetime(1995, 2, 22, 0, 0)}, {'Name': 'one', 'Gender': 'two', 'DOB': 'three'}]
    if new_sheet_name is not None:
      self.create_sheet(new_sheet_name)
      self.set_active_sheet(new_sheet_name)
    # Add headings
    headings = list(data[0].keys())
    self.working_sheet.append(headings) 
    # Add data
    for d in data:
      values = list(d.values())
      self.working_sheet.append(values) 
    self.save_file()

try:
  excel= HandleExcel('C:\Z Completed Projects\\flask\model\ExcelFile.xlsx')
  # excel.get_value_of_cell('B3')
  # excel.set_value_of_cell('B3', 'new value')
  # excel.get_value_of_cell('B3')

  # print(excel.get_active_sheet())
  # print(excel.set_active_sheet('Sheet2'))
  # print(excel.get_active_sheet())

  # print("SHEETS", excel.create_sheet('Sheet2'))
  # print("SHEETS", excel.get_sheet_names())

  # excel.append_data(['one', 'two', 'three', 'four'])
  # print("DATA", excel.get_data(5,5))

  # data =  [{'Name': 'MEOW', 'Gender': 'Male', 'DOB': datetime.datetime(1998, 2, 22, 0, 0)}, {'Name': 'NHAU', 'Gender': 'new value', 'DOB': datetime.datetime(1995, 2, 22, 0, 0)}, {'Name': 'one', 'Gender': 'two', 'DOB': 'three'}]
  # excel.insert_data(data, "TEST")

except Exception as e:
  print("MAIN ERROR", e)

